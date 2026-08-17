#!/usr/bin/env python3
"""Compare a GIZMo DCS intake workbook with the generated OPC UA contract."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


TYPE_MAP = {
    "Boolean": "Boolean",
    "DateTime": "DateTime",
    "Double": "Double",
    "Float": "Float",
    "Integer": "UInt32",
    "Long": "UInt64",
    "String": "String",
}
UNIT_ALIASES = {
    ("°C", "Cel"),
}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Validate selected XLSX tags against gizmo-opcua-contract.json"
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--contract",
        type=Path,
        default=Path("schema/gizmo-opcua-contract.json"),
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    contract = json.loads(args.contract.read_text(encoding="utf-8"))
    entries = {
        item["node_id"]: ("Variable", item) for item in contract["variables"]
    }
    entries.update(
        {item["node_id"]: ("Method", item) for item in contract["methods"]}
    )
    sheet = load_workbook(
        args.workbook,
        data_only=True,
        read_only=True,
    )["Tag List"]
    errors: list[str] = []
    seen: set[str] = set()
    selected = 0
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), 2
    ):
        subsystem, name = row[0], row[1]
        if not name:
            continue
        selected += 1
        address = row[3]
        if address in seen:
            errors.append(f"row {row_number}: duplicate NodeId {address}")
        seen.add(address)
        found = entries.get(address)
        if found is None:
            errors.append(f"row {row_number}: missing {subsystem}/{name}: {address}")
            continue
        node_class, item = found
        if subsystem == "Operation":
            if node_class != "Method":
                errors.append(
                    f"row {row_number}: {name} is {node_class}, expected Method"
                )
            if row[22] != "Write-Only":
                errors.append(
                    f"row {row_number}: method {name} is not marked Write-Only"
                )
            continue
        if node_class != "Variable":
            errors.append(
                f"row {row_number}: {name} is {node_class}, expected Variable"
            )
            continue
        expected_type = TYPE_MAP.get(str(row[4]))
        if item["data_type"] != expected_type:
            errors.append(
                f"row {row_number}: {name} type {item['data_type']} != "
                f"XLSX {row[4]} ({expected_type})"
            )
        expected_access = {
            "Read-Only": "ReadOnly",
            "Read-Write": "ReadWrite",
        }.get(row[22])
        if item["access"] != expected_access:
            errors.append(
                f"row {row_number}: {name} access {item['access']} != "
                f"XLSX {row[22]}"
            )
        requested_unit = row[5]
        actual_unit = (item.get("engineering_unit") or {}).get("symbol")
        if (
            requested_unit not in (None, "N/A")
            and requested_unit != actual_unit
            and (requested_unit, actual_unit) not in UNIT_ALIASES
        ):
            errors.append(
                f"row {row_number}: {name} unit {actual_unit!r} != "
                f"XLSX {requested_unit!r}"
            )

    if errors:
        print("\n".join(errors))
        print(
            f"FAILED: {selected - len(errors)}/{selected} selected rows "
            "passed without a reported mismatch"
        )
        return 1
    print(
        f"PASS: {selected}/{selected} selected XLSX rows are present with "
        f"matching class, datatype, access, and units in model "
        f"{contract['model_version']} ({contract['contract_sha256']})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
